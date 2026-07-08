import datetime
import re
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from ..db import get_conn

router = APIRouter()

_ISO_DATE_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
_DMY_DATE_RE = re.compile(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$")
_TIME_RE = re.compile(r"^(\d{1,2}):(\d{2})")
_VALID_STATUSES = {"pending", "corrected", "needs_review"}

HEADER_ALIASES = {
    "date": "date",
    "time": "time",
    "name": "name",
    "sender": "name",
    "sender name": "name",
    "amount": "amount",
    "notes": "notes",
    "note": "notes",
    "status": "status",
    "type": "type",
}


def _format_date_display(iso: Optional[str]) -> str:
    if not iso:
        return ""
    m = _ISO_DATE_RE.match(iso)
    if not m:
        return iso
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"


def _safe_filename_part(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", text).strip() or "export"


@router.get("/export")
def export_transactions(
    batch_id: Optional[str] = "unsaved",
    draft_slot: Optional[int] = None,
    type: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    clauses = []
    params: list = []

    batch_name = None
    if batch_id is None or batch_id == "unsaved":
        clauses.append("batch_id IS NULL")
        if draft_slot is not None:
            clauses.append("draft_slot = ?")
            params.append(draft_slot)
    elif batch_id != "all":
        with get_conn() as conn:
            batch_row = conn.execute("SELECT name FROM batches WHERE id = ?", (int(batch_id),)).fetchone()
        if not batch_row:
            raise HTTPException(404, "Batch not found")
        batch_name = batch_row["name"]
        clauses.append("batch_id = ?")
        params.append(int(batch_id))

    if type:
        clauses.append("transaction_type = ?")
        params.append(type)
    if status:
        clauses.append("status = ?")
        params.append(status)
    if search:
        clauses.append("sender_name LIKE ?")
        params.append(f"%{search}%")
    if date_from:
        clauses.append("transaction_date >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("transaction_date <= ?")
        params.append(date_to)

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    query = f"SELECT * FROM transactions {where} ORDER BY transaction_date, transaction_time"

    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(query, params).fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Time", "Name", "Amount", "Notes", "Status", "Type"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    total = 0.0
    for row in rows:
        amount = row.get("amount")
        ws.append([
            _format_date_display(row.get("transaction_date")),
            row.get("transaction_time") or "",
            row.get("sender_name") or "",
            amount if amount is not None else "",
            row.get("notes") or "",
            "needs review" if row.get("status") == "needs_review" else (row.get("status") or ""),
            (row.get("transaction_type") or "income").capitalize(),
        ])
        total += float(amount) if amount is not None else 0.0

    ws.append([])
    ws.append(["", "", "Total", total, "", "", ""])
    last_row = ws.max_row
    ws.cell(row=last_row, column=3).font = Font(bold=True)
    ws.cell(row=last_row, column=4).font = Font(bold=True)

    widths = {"A": 12, "B": 8, "C": 28, "D": 12, "E": 32, "F": 14, "G": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    type_label = (type or "income").capitalize()
    if batch_name:
        filename = f"{type_label} - {batch_name}.xlsx"
    elif batch_id == "all":
        filename = f"{type_label} - All.xlsx"
    else:
        filename = f"{type_label} - Unsaved.xlsx"
    filename = _safe_filename_part(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_date_cell(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return f"{value.year:04d}-{value.month:02d}-{value.day:02d}"
    text = str(value).strip()
    if not text:
        return None
    if _ISO_DATE_RE.match(text):
        return text
    m = _DMY_DATE_RE.match(text)
    if m:
        day, month, year = m.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    return None


def _parse_time_cell(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.time)):
        return f"{value.hour:02d}:{value.minute:02d}"
    text = str(value).strip()
    if not text:
        return None
    m = _TIME_RE.match(text)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return None


def _parse_amount_cell(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_status_cell(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().lower().replace(" ", "_")
    return text if text in _VALID_STATUSES else None


def _parse_type_cell(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().lower()
    return text if text in ("income", "expense") else None


@router.post("/import")
async def import_transactions(file: UploadFile, draft_slot: int = Form(1), type: str = Form("income")):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Please upload an .xlsx file")

    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(400, "Could not read this file — make sure it's a valid, unprotected .xlsx file")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "The file is empty")

    col_map = {}
    for idx, header in enumerate(header_row or []):
        if header is None:
            continue
        key = HEADER_ALIASES.get(str(header).strip().lower())
        if key and key not in col_map:
            col_map[key] = idx

    if "amount" not in col_map:
        raise HTTPException(
            400,
            "Couldn't find an 'Amount' column — the first row needs headers like "
            "Date, Time, Name, Amount, Notes, Status, Type",
        )

    def cell(row, key):
        idx = col_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    imported = 0
    skipped = 0
    with get_conn() as conn:
        for row in rows_iter:
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue

            name_val = cell(row, "name")
            if name_val is not None and str(name_val).strip().lower() == "total":
                continue  # the summary row from our own exported files

            amount = _parse_amount_cell(cell(row, "amount"))
            if amount is None:
                skipped += 1
                continue

            transaction_date = _parse_date_cell(cell(row, "date"))
            transaction_time = _parse_time_cell(cell(row, "time"))
            sender_name = str(name_val).strip() if name_val is not None else None
            notes_val = cell(row, "notes")
            notes = str(notes_val).strip() if notes_val is not None else None
            row_status = _parse_status_cell(cell(row, "status")) or "corrected"
            row_type = _parse_type_cell(cell(row, "type")) or type

            conn.execute(
                """INSERT INTO transactions
                    (batch_id, draft_slot, transaction_type, sender_name, transaction_date,
                     transaction_time, amount, notes, status)
                   VALUES (NULL, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (draft_slot, row_type, sender_name, transaction_date, transaction_time, amount, notes, row_status),
            )
            imported += 1

    return {"imported": imported, "skipped": skipped}
